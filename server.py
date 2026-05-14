import io
import json
import os
import queue
import re
import threading
import time
import uuid

from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Histórico persistente ──────────────────────────────────────────────────────
_HIST_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "historico.json")


def _chave_lead(lead: dict) -> str:
    nome = re.sub(r"\s+", " ", lead.get("nome", "")).lower().strip()
    tel = re.sub(r"\D", "", lead.get("telefone", ""))
    if tel:
        return f"{nome}|{tel}"
    site = re.sub(r"^https?://(www\.)?", "", lead.get("site", "").lower().rstrip("/"))
    return f"{nome}|{site}"


def carregar_historico() -> dict:
    if os.path.exists(_HIST_PATH):
        try:
            with open(_HIST_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"chaves": [], "total": 0}


def salvar_historico(hist: dict):
    with open(_HIST_PATH, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)


# ── Excel ──────────────────────────────────────────────────────────────────────
def gerar_excel(leads, nicho, local) -> bytes:
    COR_CAB = "1F4E79"
    COR_PAR = "D9E1F2"

    def borda():
        lado = Side(style="thin", color="AAAAAA")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    cols = ["#", "Nome da Empresa", "Telefone", "Email", "Instagram", "Site"]
    widths = [5, 40, 20, 38, 28, 45]

    for c, (t, w) in enumerate(zip(cols, widths), 1):
        cel = ws.cell(row=1, column=c, value=t)
        cel.font = Font(bold=True, color="FFFFFF", size=11)
        cel.fill = PatternFill("solid", fgColor=COR_CAB)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = borda()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    for idx, lead in enumerate(leads, 1):
        row = idx + 1
        vals = [idx, lead.get("nome", ""), lead.get("telefone", ""),
                lead.get("email", ""), lead.get("instagram", ""), lead.get("site", "")]
        fill = PatternFill("solid", fgColor=COR_PAR) if idx % 2 == 0 else None
        for c, v in enumerate(vals, 1):
            cel = ws.cell(row=row, column=c, value=v)
            cel.alignment = Alignment(vertical="center")
            cel.border = borda()
            if fill:
                cel.fill = fill

    ws.freeze_panes = "A2"
    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    m = ws["A1"]
    m.value = f"Leads — {nicho.title()} em {local.title()} | {len(leads)} empresas"
    m.font = Font(bold=True, size=13, color="1F4E79")
    m.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Worker (scraping) ──────────────────────────────────────────────────────────
def worker(nicho, local, max_leads, log_q, result_q):
    try:
        import requests
        from bs4 import BeautifulSoup
        import undetected_chromedriver as uc
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from urllib.parse import urlparse
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        HEADERS = {"User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )}
        EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
        IGNORE = {"exemplo.com", "example.com", "seudominio.com", "wixpress.com", "sentry.io"}
        INSTA_RE = re.compile(r'instagram\.com/([A-Za-z0-9._]{2,30})/?(?:["\s?#]|$)')
        INSTA_SKIP = {"p", "reel", "reels", "explore", "accounts", "tv", "stories", "direct",
                      "share", "about", "legal", "privacy", "security", "blog", "press",
                      "api", "developer", "_n", "sharedfiles"}

        def log(msg):
            log_q.put(msg)

        def limpar_tel(t):
            return re.sub(r"[^\d+()\s\-]", "", t).strip()

        def fmt_url(u):
            if not u:
                return ""
            return u if u.startswith(("http://", "https://")) else "https://" + u

        def extrair_instagram(html):
            for m in INSTA_RE.finditer(html):
                handle = m.group(1).lower()
                if handle not in INSTA_SKIP and not handle.startswith("_"):
                    return "@" + handle
            return ""

        def buscar_contato(url):
            emails = set()
            instagram = ""
            url = fmt_url(url)
            if not url:
                return "", ""
            try:
                r = requests.get(url, headers=HEADERS, timeout=8, verify=False, allow_redirects=True)
                soup = BeautifulSoup(r.text, "html.parser")
                for tag in soup.find_all("a", href=True):
                    h = tag["href"]
                    if "instagram.com/" in h and not instagram:
                        instagram = extrair_instagram(h + '"')
                    if h.startswith("mailto:"):
                        e = h.replace("mailto:", "").split("?")[0].strip().lower()
                        d = e.split("@")[-1]
                        if "@" in e and d not in IGNORE:
                            emails.add(e)
                if not instagram:
                    instagram = extrair_instagram(r.text)
                if not emails:
                    for m in EMAIL_RE.findall(r.text):
                        d = m.split("@")[-1].lower()
                        if d not in IGNORE:
                            emails.add(m.lower())
                            break
                if not emails or not instagram:
                    base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
                    for slug in ["/contato", "/contact", "/fale-conosco", "/sobre"]:
                        try:
                            r2 = requests.get(base + slug, headers=HEADERS, timeout=8, verify=False)
                            if not emails:
                                for m in EMAIL_RE.findall(r2.text):
                                    d = m.split("@")[-1].lower()
                                    if d not in IGNORE:
                                        emails.add(m.lower())
                            if not instagram:
                                instagram = extrair_instagram(r2.text)
                            if emails and instagram:
                                break
                        except Exception:
                            pass
            except Exception:
                pass
            return (list(emails)[0] if emails else ""), instagram

        log("🚀 Iniciando navegador...")

        def _chrome_version():
            import subprocess
            import re as _re
            try:
                import winreg
                for root in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
                    for sub in (
                        r"SOFTWARE\Google\Chrome\BLBeacon",
                        r"SOFTWARE\Wow6432Node\Google\Chrome\BLBeacon",
                    ):
                        try:
                            key = winreg.OpenKey(root, sub)
                            val, _ = winreg.QueryValueEx(key, "version")
                            winreg.CloseKey(key)
                            m = _re.match(r"(\d+)", str(val))
                            if m:
                                return int(m.group(1))
                        except OSError:
                            pass
            except ImportError:
                pass
            paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                "/usr/bin/google-chrome",
                "/usr/bin/chromium-browser",
                "/usr/bin/chromium",
            ]
            for p in paths:
                try:
                    out = subprocess.check_output([p, "--version"], stderr=subprocess.DEVNULL, text=True, timeout=5)
                    m = _re.search(r"(\d+)\.", out)
                    if m:
                        return int(m.group(1))
                except Exception:
                    pass
            return None

        def _build_opts():
            o = uc.ChromeOptions()
            o.add_argument("--no-sandbox")
            o.add_argument("--disable-dev-shm-usage")
            o.add_argument("--disable-gpu")
            o.add_argument("--disable-software-rasterizer")
            o.add_argument("--disable-extensions")
            o.add_argument("--disable-plugins")
            o.add_argument("--blink-settings=imagesEnabled=false")
            o.add_argument("--disable-background-networking")
            o.add_argument("--disable-default-apps")
            o.add_argument("--disable-sync")
            o.add_argument("--disable-translate")
            o.add_argument("--mute-audio")
            o.add_argument("--no-first-run")
            o.add_argument("--disable-background-timer-throttling")
            o.add_argument("--disable-renderer-backgrounding")
            o.add_argument("--disable-blink-features=AutomationControlled")
            o.add_argument("--js-flags=--max-old-space-size=256")
            o.add_argument("--window-size=1024,600")
            o.add_argument("--lang=pt-BR")
            o.add_argument("--headless=new")
            return o

        ver = _chrome_version()
        log(f"   Versão do Chrome detectada: {ver or 'desconhecida'}")

        driver = None
        versions_to_try = [ver, ver - 1, ver + 1] if ver else [None]
        for v in versions_to_try:
            kw = {"version_main": v} if v is not None else {}
            try:
                driver = uc.Chrome(options=_build_opts(), **kw)
                log(f"   Chrome iniciado com version_main={v or 'auto'}")
                break
            except Exception:
                log(f"   version_main={v} falhou, tentando próxima...")
                driver = None

        if driver is None:
            raise RuntimeError(
                f"Não foi possível iniciar o Chrome (versão {ver}). "
                "Atualize: pip install -U undetected-chromedriver"
            )

        driver.set_window_size(1280, 720)
        query = f"{nicho} em {local}"
        log(f"🗺️  Buscando: {query}")
        driver.get(f"https://www.google.com/maps/search/{query.replace(' ', '+')}")
        time.sleep(3)

        n_scrolls = max(10, (max_leads // 5) + 5)
        log("⏬ Carregando resultados...")

        try:
            painel = driver.find_element(
                By.XPATH,
                '//div[contains(@aria-label,"Resultados") or contains(@aria-label,"Results")][@role="feed"]'
            )
        except Exception:
            painel = None

        for _ in range(n_scrolls):
            try:
                if painel:
                    driver.execute_script("arguments[0].scrollBy(0, 800);", painel)
                else:
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
                time.sleep(2)
            except Exception:
                break

        cards = []
        for sel in ['a[href*="/maps/place/"]', 'div[role="article"] a[href*="maps"]']:
            cards = driver.find_elements(By.CSS_SELECTOR, sel)
            if cards:
                break

        log(f"📋 {len(cards)} empresas encontradas. Coletando detalhes...")

        hist = carregar_historico()
        chaves_set = set(hist.get("chaves", []))
        leads = []
        pulados = 0
        visitados = set()
        wait_short = WebDriverWait(driver, 5)

        for i, card in enumerate(cards):
            if len(leads) >= max_leads:
                break
            try:
                href = card.get_attribute("href") or ""
                if href in visitados:
                    continue
                visitados.add(href)

                log(f"🔍 [{i+1}/{len(cards)}] Abrindo empresa...")
                driver.execute_script("arguments[0].click();", card)
                time.sleep(3)

                lead = {"nome": "", "telefone": "", "site": "", "email": "", "instagram": ""}

                for sel in [
                    'h1[class*="fontHeadlineLarge"]', 'h1[class*="DUwDvf"]',
                    'h1.DUwDvf', 'div[role="main"] h1', 'h1',
                ]:
                    try:
                        el = wait_short.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                        nome = el.text.strip()
                        if nome:
                            lead["nome"] = nome
                            break
                    except Exception:
                        pass

                if not lead["nome"]:
                    continue

                try:
                    spans = driver.find_elements(By.XPATH,
                        '//button[@data-tooltip="Copiar número de telefone"]'
                        '//div[contains(@class,"fontBodyMedium")]')
                    if not spans:
                        spans = driver.find_elements(By.XPATH,
                            '//*[contains(@aria-label,"Ligue") or contains(@aria-label,"Phone")]')
                    for el in spans:
                        t = limpar_tel(el.get_attribute("aria-label") or el.text)
                        if re.search(r"\d{4,}", t):
                            lead["telefone"] = t
                            break
                except Exception:
                    pass

                if not lead["telefone"]:
                    try:
                        for btn in driver.find_elements(By.XPATH, '//button[@aria-label]'):
                            lbl = btn.get_attribute("aria-label") or ""
                            if re.search(r"\(\d{2}\)|\+\d{2}|\d{8,}", lbl):
                                lead["telefone"] = limpar_tel(lbl)
                                break
                    except Exception:
                        pass

                try:
                    sb = driver.find_element(By.XPATH,
                        '//a[@data-tooltip="Abrir site" or @data-item-id="authority"]')
                    lead["site"] = sb.get_attribute("href") or ""
                except Exception:
                    pass

                if lead["site"]:
                    log(f"   [{len(leads)+1}] {lead['nome'][:40]} — buscando contato...")
                    lead["email"], lead["instagram"] = buscar_contato(lead["site"])
                else:
                    log(f"   [{len(leads)+1}] {lead['nome'][:40]} — sem site")

                chave = _chave_lead(lead)
                if chave in chaves_set:
                    pulados += 1
                    log(f"   ⏭️  {lead['nome'][:40]} — já coletado antes, pulando")
                    continue

                chaves_set.add(chave)
                leads.append(lead)
                log(f"       ✓ Tel: {lead['telefone'] or '—'} | Email: {lead['email'] or '—'} | Insta: {lead['instagram'] or '—'}")
                result_q.put(("lead", lead))

            except Exception as e:
                log(f"   [!] Erro no card {i+1}: {e}")

        try:
            driver.quit()
        except Exception:
            pass

        hist["chaves"] = list(chaves_set)
        hist["total"] = len(chaves_set)
        salvar_historico(hist)
        if pulados:
            log(f"💾 {pulados} empresa(s) já conhecida(s) foram ignoradas.")
        log(f"💾 Histórico salvo: {len(chaves_set)} empresas únicas no total.")
        result_q.put(("done", leads))

    except Exception as e:
        log_q.put(f"❌ Erro fatal: {e}")
        result_q.put(("done", []))


# ── Flask app ──────────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app, resources={r"/*": {"origins": "*"}})


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, DELETE, OPTIONS"
    return response


@app.route("/api/<path:path>", methods=["OPTIONS"])
def options_handler(path):
    return "", 204

_jobs: dict = {}


@app.route("/")
def index():
    return app.send_static_file("index.html")


@app.route("/api/historico", methods=["GET"])
def get_historico():
    hist = carregar_historico()
    return jsonify({"total": hist.get("total", len(hist.get("chaves", [])))})


@app.route("/api/historico", methods=["DELETE"])
def del_historico():
    salvar_historico({"chaves": [], "total": 0})
    return jsonify({"ok": True})


@app.route("/api/buscar", methods=["POST"])
def buscar():
    data = request.get_json(force=True) or {}
    nicho = data.get("nicho", "").strip()
    local = data.get("local", "São Paulo").strip()
    max_leads = max(1, min(200, int(data.get("max_leads", 30))))

    if not nicho:
        return jsonify({"error": "Informe o nicho"}), 400

    job_id = str(uuid.uuid4())
    log_q = queue.Queue()
    result_q = queue.Queue()

    _jobs[job_id] = {
        "log_q": log_q,
        "logs": [],
        "leads": [],
        "done": False,
        "excel": None,
        "nicho": nicho,
        "local": local,
    }

    def _process():
        while True:
            # Drain log queue into persistent list
            while not log_q.empty():
                try:
                    _jobs[job_id]["logs"].append(log_q.get_nowait())
                except queue.Empty:
                    break
            try:
                kind, payload = result_q.get(timeout=0.3)
                if kind == "lead":
                    _jobs[job_id]["leads"].append(payload)
                elif kind == "done":
                    # Drain remaining logs before marking done
                    while not log_q.empty():
                        try:
                            _jobs[job_id]["logs"].append(log_q.get_nowait())
                        except queue.Empty:
                            break
                    _jobs[job_id]["done"] = True
                    if _jobs[job_id]["leads"]:
                        _jobs[job_id]["excel"] = gerar_excel(_jobs[job_id]["leads"], nicho, local)
                    break
            except queue.Empty:
                pass

    threading.Thread(target=worker, args=(nicho, local, max_leads, log_q, result_q), daemon=True).start()
    threading.Thread(target=_process, daemon=True).start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def status(job_id):
    if job_id not in _jobs:
        return jsonify({"error": "Job não encontrado"}), 404
    job = _jobs[job_id]
    log_offset  = int(request.args.get("log_offset",  0))
    lead_offset = int(request.args.get("lead_offset", 0))
    return jsonify({
        "logs":        job["logs"][log_offset:],
        "leads":       job["leads"][lead_offset:],
        "total_leads": len(job["leads"]),
        "total_logs":  len(job["logs"]),
        "done":        job["done"],
        "has_excel":   job["excel"] is not None,
    })


@app.route("/api/stream/<job_id>")
def stream(job_id):
    if job_id not in _jobs:
        return jsonify({"error": "Job não encontrado"}), 404

    def generate():
        job = _jobs[job_id]
        sent = 0
        while True:
            while not job["log_q"].empty():
                msg = job["log_q"].get()
                yield f"data: {json.dumps({'t': 'log', 'msg': msg})}\n\n"

            current = len(job["leads"])
            if current > sent:
                new_leads = job["leads"][sent:current]
                yield f"data: {json.dumps({'t': 'leads', 'leads': new_leads, 'total': current})}\n\n"
                sent = current

            if job["done"]:
                yield f"data: {json.dumps({'t': 'done', 'total': current, 'has_excel': job['excel'] is not None})}\n\n"
                break

            time.sleep(0.8)

    return Response(generate(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
    })


@app.route("/api/download/<job_id>")
def download(job_id):
    if job_id not in _jobs or not _jobs[job_id]["excel"]:
        return jsonify({"error": "Excel não disponível"}), 404
    job = _jobs[job_id]
    fname = f"leads_{job['nicho'].replace(' ', '_')}_{job['local'].replace(' ', '_')}.xlsx"
    return send_file(
        io.BytesIO(job["excel"]),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
