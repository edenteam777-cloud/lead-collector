import streamlit as st
import threading
import queue
import io
import time
import re
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Histórico persistente ──────────────────────────────────────────────────────
_HIST_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "historico.json")

def _chave_lead(lead: dict) -> str:
    nome = re.sub(r"\s+", " ", lead.get("nome", "")).lower().strip()
    tel  = re.sub(r"\D", "", lead.get("telefone", ""))
    if tel:
        return f"{nome}|{tel}"
    site = re.sub(r"^https?://(www\.)?", "", lead.get("site", "").lower().rstrip("/"))
    return f"{nome}|{site}"

def carregar_historico() -> dict:
    if os.path.exists(_HIST_PATH):
        try:
            with open(_HIST_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"chaves": [], "total": 0}

def salvar_historico(hist: dict):
    with open(_HIST_PATH, "w", encoding="utf-8") as f:
        json.dump(hist, f, ensure_ascii=False, indent=2)

# ── Colunas na ordem pedida ────────────────────────────────────────────────────
_KEYS  = ["nome", "nicho", "telefone", "email", "instagram", "cidade",
          "ticket_medio", "tempo_empresa", "porte", "cnpj", "info_empresa"]
_HEADS = ["#", "Nome", "Nicho", "Telefone", "E-mail", "Instagram",
          "Cidade", "Ticket Médio", "Tempo de Empresa", "Porte", "CNPJ", "Informações"]

def _leads_html_table(leads: list) -> str:
    rows = []
    for i, lead in enumerate(leads, 1):
        cells = "".join(
            f"<td style='padding:4px 10px;border:1px solid #ddd;white-space:nowrap;max-width:260px;"
            f"overflow:hidden;text-overflow:ellipsis'>{lead.get(k, '') or ''}</td>"
            for k in _KEYS
        )
        bg = "#f5f8ff" if i % 2 == 0 else "white"
        rows.append(
            f"<tr style='background:{bg}'>"
            f"<td style='padding:4px 8px;border:1px solid #ddd;color:#888;text-align:right'>{i}</td>"
            f"{cells}</tr>"
        )
    header = "".join(
        f"<th style='padding:6px 10px;background:#1F4E79;color:white;border:1px solid #1a4268;"
        f"white-space:nowrap'>{h}</th>"
        for h in _HEADS
    )
    return (
        "<div style='overflow-x:auto;max-height:420px;overflow-y:auto'>"
        f"<table style='border-collapse:collapse;font-size:13px;width:100%'>"
        f"<thead><tr>{header}</tr></thead><tbody>{''.join(rows)}</tbody></table></div>"
    )

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Coletor de Leads", page_icon="🎯", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .hero {
        background: linear-gradient(135deg, #1F4E79 0%, #2E86C1 100%);
        border-radius: 16px; padding: 2.5rem 2rem 2rem; text-align: center;
        margin-bottom: 2rem; box-shadow: 0 4px 24px rgba(31,78,121,0.18);
    }
    .hero h1 { color:#fff; font-size:2.4rem; font-weight:700; margin:0; }
    .hero p  { color:#D6EAF8; font-size:1rem; margin-top:.5rem; }
    .card { background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1);
            border-radius:12px; padding:1.5rem 1.8rem; margin-bottom:1.4rem; }
    .stat-box { background:#1F4E79; color:#fff; border-radius:10px; padding:1rem; text-align:center; }
    .stat-box .num { font-size:2rem; font-weight:700; }
    .stat-box .lbl { font-size:.8rem; opacity:.8; }
    .log-box { background:#0d1117; color:#58a6ff; border-radius:10px; padding:1rem 1.2rem;
               font-family:monospace; font-size:.85rem; max-height:260px; overflow-y:auto;
               white-space:pre-wrap; line-height:1.6; }
    div[data-testid="stButton"] > button {
        width:100%; background:linear-gradient(90deg,#1F4E79,#2E86C1); color:white;
        border:none; border-radius:8px; padding:.75rem 1.5rem; font-size:1rem;
        font-weight:600; cursor:pointer; transition:opacity .2s;
    }
    div[data-testid="stButton"] > button:hover { opacity:.88; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
    <h1>🎯 Coletor de Leads</h1>
    <p>Extrai empresas do Google Maps com dados completos — exporta para Excel em segundos.</p>
</div>
""", unsafe_allow_html=True)

# ── Session state ──────────────────────────────────────────────────────────────
for key, val in {"running": False, "leads": [], "logs": [], "done": False, "excel_bytes": None}.items():
    if key not in st.session_state:
        st.session_state[key] = val

_hist_info  = carregar_historico()
_total_hist = _hist_info.get("total", len(_hist_info.get("chaves", [])))
if _total_hist > 0:
    col_h1, col_h2 = st.columns([3, 1])
    with col_h1:
        st.info(f"💾 **Memória ativa:** {_total_hist} empresa(s) já coletadas — duplicatas ignoradas automaticamente.")
    with col_h2:
        if st.button("🗑️ Limpar memória"):
            salvar_historico({"chaves": [], "total": 0})
            st.success("Histórico apagado!")
            st.rerun()

# ── Excel ──────────────────────────────────────────────────────────────────────
def gerar_excel(leads, nicho, local) -> bytes:
    COR_CAB = "1F4E79"
    COR_PAR = "D9E1F2"

    def borda():
        lado = Side(style="thin", color="AAAAAA")
        return Border(left=lado, right=lado, top=lado, bottom=lado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    cols   = ["#", "Nome da Empresa", "Nicho", "Telefone", "Email",
              "Instagram", "Cidade", "Ticket Médio", "Tempo de Empresa",
              "Porte", "CNPJ", "Informações"]
    widths = [5, 32, 18, 18, 32, 22, 18, 14, 18, 12, 20, 50]

    for c, (t, w) in enumerate(zip(cols, widths), 1):
        cel = ws.cell(row=1, column=c, value=t)
        cel.font      = Font(bold=True, color="FFFFFF", size=11)
        cel.fill      = PatternFill("solid", fgColor=COR_CAB)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border    = borda()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    for idx, lead in enumerate(leads, 1):
        row  = idx + 1
        vals = [idx,
                lead.get("nome", ""),
                lead.get("nicho", nicho),
                lead.get("telefone", ""),
                lead.get("email", ""),
                lead.get("instagram", ""),
                lead.get("cidade", ""),
                lead.get("ticket_medio", ""),
                lead.get("tempo_empresa", ""),
                lead.get("porte", ""),
                lead.get("cnpj", ""),
                lead.get("info_empresa", "")]
        fill = PatternFill("solid", fgColor=COR_PAR) if idx % 2 == 0 else None
        for c, v in enumerate(vals, 1):
            cel = ws.cell(row=row, column=c, value=v)
            cel.alignment = Alignment(vertical="center", wrap_text=(c == 10))
            cel.border    = borda()
            if fill:
                cel.fill = fill

    ws.freeze_panes = "A2"
    ws.insert_rows(1)
    ws.merge_cells("A1:L1")
    m = ws["A1"]
    m.value     = f"Leads — {nicho.title()} em {local.title()} | {len(leads)} empresas"
    m.font      = Font(bold=True, size=13, color="1F4E79")
    m.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Worker ─────────────────────────────────────────────────────────────────────
def worker(nicho, local, max_leads, log_q, result_q):
    try:
        import requests
        from bs4 import BeautifulSoup
        import undetected_chromedriver as uc
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from urllib.parse import urlparse
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        HEADERS = {"User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )}
        EMAIL_RE  = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
        IGNORE    = {"exemplo.com", "example.com", "seudominio.com", "wixpress.com", "sentry.io"}
        INSTA_RE  = re.compile(r'instagram\.com/([A-Za-z0-9._]{2,30})/?(?:["\s?#]|$)')
        INSTA_SKIP = {"p","reel","reels","explore","accounts","tv","stories","direct","share",
                      "about","legal","privacy","security","blog","press","api","developer","_n","sharedfiles"}
        ANO_RE    = re.compile(r'(?:desde|fundad[ao]|criado?|estabelecid[ao]|aberto?[\s]+desde|inaugurado?)[:\s]+(\d{4})', re.I)
        CNPJ_RE   = re.compile(r'\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\s\/]?\d{4}[\s\-]?\d{2}')

        def log(msg): log_q.put(msg)

        def limpar_tel(t): return re.sub(r"[^\d+()\s\-]", "", t).strip()

        def fmt_url(u):
            if not u: return ""
            return u if u.startswith(("http://", "https://")) else "https://" + u

        def extrair_instagram(html):
            for m in INSTA_RE.finditer(html):
                handle = m.group(1).lower()
                if handle not in INSTA_SKIP and not handle.startswith("_"):
                    return "@" + handle
            return ""

        def extrair_cidade(endereco: str) -> str:
            if not endereco: return ""
            e = re.sub(r'^Endereço:\s*|^Address:\s*', '', endereco, flags=re.I)
            partes = [p.strip() for p in re.split(r'[,\-]', e)]
            skip = re.compile(r'^\d|^(rua|av\.|avenida|r\.|al\.|alameda|praça|travessa|estrada|rod\.)|^\d{4,}', re.I)
            estado = re.compile(r'^[A-Z]{2}$')
            cep    = re.compile(r'\d{4,5}-?\d{3}')
            paises = {'brasil', 'brazil', 'portugal', 'españa', 'spain', 'france', 'italia'}
            candidatos = [p for p in partes if p and len(p) >= 3
                          and not skip.match(p) and not estado.match(p)
                          and not cep.search(p) and p.lower() not in paises]
            if len(candidatos) >= 2:
                return candidatos[-2]
            return candidatos[-1] if candidatos else ""

        def calcular_tempo(ano: int) -> str:
            anos = 2025 - ano
            if anos <= 0: return ""
            return f"{anos} ano{'s' if anos != 1 else ''} (desde {ano})"

        def extrair_cnpj(html: str) -> str:
            m = CNPJ_RE.search(html)
            if not m: return ""
            digits = re.sub(r"\D", "", m.group())
            return digits if len(digits) == 14 else ""

        def buscar_cnpj(cnpj: str) -> dict:
            """Consulta BrasilAPI e retorna dict com campos enriquecidos."""
            resultado = {}
            try:
                r = requests.get(
                    f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}",
                    headers=HEADERS, timeout=10
                )
                if r.status_code != 200:
                    return resultado
                d = r.json()

                # Tempo de empresa (data_abertura: "YYYY-MM-DD" ou "DD/MM/YYYY")
                abertura = d.get("data_abertura", "")
                if abertura:
                    try:
                        if "-" in abertura:
                            ano = int(abertura.split("-")[0])
                        else:
                            ano = int(abertura.split("/")[-1])
                        if 1900 <= ano <= 2025:
                            resultado["tempo_empresa"] = calcular_tempo(ano)
                    except Exception:
                        pass

                # Porte
                porte_raw = d.get("porte", "") or ""
                porte_map = {
                    "MEI": "MEI", "ME": "ME (Microempresa)",
                    "EPP": "EPP (Pequeno Porte)",
                    "DEMAIS": "Médio/Grande Porte",
                }
                resultado["porte"] = porte_map.get(porte_raw.upper(), porte_raw.title())

                # Atividade principal
                atividades = d.get("descricao_atividade_principal") or d.get("cnaes_secundarios") or []
                if atividades and isinstance(atividades, list):
                    resultado["atividade"] = atividades[0].get("text", "").strip()

                # Cidade via CNPJ (complemento se não encontrada)
                municipio = d.get("municipio", "")
                uf        = d.get("uf", "")
                if municipio:
                    resultado["cidade_cnpj"] = f"{municipio.title()} - {uf}" if uf else municipio.title()

                # CNPJ formatado
                resultado["cnpj_fmt"] = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

            except Exception:
                pass
            return resultado

        def pesquisar_ddg(nome: str, cidade: str) -> dict:
            """Busca '{nome} {cidade} CNPJ' no DuckDuckGo e extrai dados."""
            res = {"cnpj": "", "email": "", "instagram": "", "info": ""}
            try:
                query = f'"{nome}" {cidade} CNPJ'
                r = requests.get(
                    "https://html.duckduckgo.com/html/",
                    params={"q": query, "kl": "br-pt"},
                    headers={**HEADERS, "Accept-Language": "pt-BR,pt;q=0.9"},
                    timeout=12
                )
                soup = BeautifulSoup(r.text, "html.parser")
                texto = soup.get_text(" ", strip=True)

                # CNPJ nos resultados
                res["cnpj"] = extrair_cnpj(texto)

                # Email
                for m in EMAIL_RE.findall(texto):
                    d = m.split("@")[-1].lower()
                    if d not in IGNORE:
                        res["email"] = m.lower(); break

                # Instagram
                res["instagram"] = extrair_instagram(r.text)

                # Info: snippet do 1º resultado relevante
                for sel in ["a.result__snippet", "div.result__snippet", ".result__body"]:
                    for tag in soup.select(sel)[:3]:
                        txt = tag.get_text(" ", strip=True)
                        if len(txt) > 40:
                            res["info"] = txt[:280]; break
                    if res["info"]: break

                # Se não achou CNPJ, tenta busca mais direta
                if not res["cnpj"]:
                    query2 = f"{nome} {cidade} CNPJ site:cnpj.biz OR site:cnpja.com OR site:casadosdados.com.br"
                    r2 = requests.get(
                        "https://html.duckduckgo.com/html/",
                        params={"q": query2, "kl": "br-pt"},
                        headers={**HEADERS, "Accept-Language": "pt-BR,pt;q=0.9"},
                        timeout=10
                    )
                    res["cnpj"] = extrair_cnpj(BeautifulSoup(r2.text, "html.parser").get_text(" "))

            except Exception:
                pass
            return res

        def buscar_site(url):
            """Retorna dict com email, instagram, tempo, info, cnpj_data."""
            result = {"email":"","instagram":"","tempo":"","info":"","cnpj_data":{}}
            url = fmt_url(url)
            if not url: return result
            try:
                r = requests.get(url, headers=HEADERS, timeout=8, verify=False, allow_redirects=True)
                html = r.text
                soup = BeautifulSoup(html, "html.parser")
                emails = set()

                for tag in soup.find_all("a", href=True):
                    h = tag["href"]
                    if "instagram.com/" in h and not result["instagram"]:
                        result["instagram"] = extrair_instagram(h + '"')
                    if h.startswith("mailto:"):
                        e = h.replace("mailto:", "").split("?")[0].strip().lower()
                        d = e.split("@")[-1]
                        if "@" in e and d not in IGNORE:
                            emails.add(e)

                if not result["instagram"]:
                    result["instagram"] = extrair_instagram(html)
                if not emails:
                    for m in EMAIL_RE.findall(html):
                        d = m.split("@")[-1].lower()
                        if d not in IGNORE:
                            emails.add(m.lower()); break

                # Tempo de empresa (site)
                m = ANO_RE.search(html)
                if m:
                    ano = int(m.group(1))
                    if 1900 <= ano <= 2025:
                        result["tempo"] = calcular_tempo(ano)

                # Info: meta description ou 1º parágrafo relevante
                meta = soup.find("meta", attrs={"name": re.compile(r"description", re.I)})
                if meta and meta.get("content", "").strip():
                    result["info"] = meta["content"].strip()[:250]
                if not result["info"]:
                    for tag in soup.find_all(["p", "h2", "h3"]):
                        txt = tag.get_text(" ", strip=True)
                        if len(txt) > 40:
                            result["info"] = txt[:250]; break

                # CNPJ no HTML principal
                cnpj = extrair_cnpj(html)

                # Páginas internas (contato/sobre) para complementar
                if not emails or not result["instagram"] or not result["tempo"] or not cnpj:
                    base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
                    for slug in ["/contato", "/contact", "/sobre", "/about", "/fale-conosco"]:
                        try:
                            r2 = requests.get(base + slug, headers=HEADERS, timeout=6, verify=False)
                            h2 = r2.text
                            if not emails:
                                for m2 in EMAIL_RE.findall(h2):
                                    d = m2.split("@")[-1].lower()
                                    if d not in IGNORE:
                                        emails.add(m2.lower())
                            if not result["instagram"]:
                                result["instagram"] = extrair_instagram(h2)
                            if not result["tempo"]:
                                m3 = ANO_RE.search(h2)
                                if m3:
                                    ano = int(m3.group(1))
                                    if 1900 <= ano <= 2025:
                                        result["tempo"] = calcular_tempo(ano)
                            if not cnpj:
                                cnpj = extrair_cnpj(h2)
                            if emails and result["instagram"] and result["tempo"] and cnpj:
                                break
                        except Exception:
                            pass

                result["email"] = list(emails)[0] if emails else ""

                # BrasilAPI: enriquece com dados oficiais da Receita Federal
                if cnpj:
                    result["cnpj_data"] = buscar_cnpj(cnpj)

            except Exception:
                pass
            return result

        # ── Chrome ──────────────────────────────────────────────────────────────
        log("🚀 Iniciando navegador...")

        def _chrome_version():
            import subprocess, re as _re
            try:
                import winreg
                for root in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
                    for sub in (r"SOFTWARE\Google\Chrome\BLBeacon",
                                r"SOFTWARE\Wow6432Node\Google\Chrome\BLBeacon"):
                        try:
                            key = winreg.OpenKey(root, sub)
                            val, _ = winreg.QueryValueEx(key, "version")
                            winreg.CloseKey(key)
                            m = _re.match(r"(\d+)", str(val))
                            if m: return int(m.group(1))
                        except OSError:
                            pass
            except ImportError:
                pass
            for p in [r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                      r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                      "/usr/bin/google-chrome", "/usr/bin/chromium-browser", "/usr/bin/chromium"]:
                try:
                    out = subprocess.check_output([p, "--version"], stderr=subprocess.DEVNULL, text=True, timeout=5)
                    m = _re.search(r"(\d+)\.", out)
                    if m: return int(m.group(1))
                except Exception:
                    pass
            return None

        def _build_opts():
            o = uc.ChromeOptions()
            o.add_argument("--no-sandbox")
            o.add_argument("--disable-dev-shm-usage")
            o.add_argument("--disable-gpu")
            o.add_argument("--disable-extensions")
            o.add_argument("--disable-background-networking")
            o.add_argument("--disable-default-apps")
            o.add_argument("--disable-sync")
            o.add_argument("--disable-translate")
            o.add_argument("--mute-audio")
            o.add_argument("--no-first-run")
            o.add_argument("--disable-blink-features=AutomationControlled")
            o.add_argument("--lang=pt-BR")
            return o

        ver = _chrome_version()
        log(f"   Chrome: {ver or 'auto-detect'}")

        driver = None
        for v in ([ver, ver - 1, ver + 1] if ver else [None]):
            kw = {"version_main": v} if v is not None else {}
            try:
                driver = uc.Chrome(options=_build_opts(), **kw)
                log(f"   Iniciado com version_main={v or 'auto'}")
                break
            except Exception:
                driver = None

        if driver is None:
            raise RuntimeError("Não foi possível iniciar o Chrome. Atualize: pip install -U undetected-chromedriver")

        driver.set_window_size(1400, 900)

        # ── Busca no Google Maps ─────────────────────────────────────────────────
        query = f"{nicho} em {local}"
        log(f"🗺️  Buscando: {query}")
        driver.get(f"https://www.google.com/maps/search/{query.replace(' ', '+')}")
        time.sleep(3)

        n_scrolls = max(15, (max_leads // 3) + 8)
        log("⏬ Carregando resultados...")

        # Detecta o painel lateral de resultados com múltiplos seletores
        painel = None
        for xpath in [
            '//div[@role="feed"]',
            '//div[contains(@aria-label,"Resultados") or contains(@aria-label,"Results")]',
            '//div[contains(@class,"m6QErb") and contains(@class,"DxyBCb")]',
            '//div[contains(@class,"m6QErb")][@aria-label]',
        ]:
            try:
                el = driver.find_element(By.XPATH, xpath)
                if el:
                    painel = el
                    break
            except Exception:
                pass

        log(f"   Painel: {'encontrado' if painel else 'fallback body'} | {n_scrolls} scrolls")

        prev_count = 0
        sem_novos  = 0
        for i in range(n_scrolls):
            try:
                if painel:
                    driver.execute_script("arguments[0].scrollTop += 900", painel)
                else:
                    driver.execute_script("window.scrollBy(0, 900)")
                time.sleep(1.8)

                # A cada 4 scrolls verifica se novos cards apareceram
                if i % 4 == 3:
                    cur = len(driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]'))
                    if cur == prev_count:
                        sem_novos += 1
                        if sem_novos >= 2:
                            log(f"   Sem novos resultados após {i+1} scrolls, parando.")
                            break
                    else:
                        sem_novos = 0
                    prev_count = cur

                    # Tenta encontrar painel se ainda não achou
                    if painel is None:
                        for xpath in ['//div[@role="feed"]',
                                      '//div[contains(@class,"m6QErb")][@aria-label]']:
                            try:
                                painel = driver.find_element(By.XPATH, xpath)
                                if painel: break
                            except Exception:
                                pass
            except Exception:
                break

        # Coleta cards únicos por href
        raw_cards = []
        for sel in [
            'div[role="feed"] a[href*="/maps/place/"]',
            'div[role="article"] a[href*="/maps/place/"]',
            'a[href*="/maps/place/"]',
        ]:
            raw_cards = driver.find_elements(By.CSS_SELECTOR, sel)
            if len(raw_cards) > 3:
                break

        seen_hrefs = set()
        cards = []
        for c in raw_cards:
            href = (c.get_attribute("href") or "").split("?")[0]
            if href and href not in seen_hrefs:
                seen_hrefs.add(href)
                cards.append(c)

        log(f"📋 {len(cards)} empresas encontradas. Coletando detalhes...")

        hist       = carregar_historico()
        chaves_set = set(hist.get("chaves", []))
        leads      = []
        pulados    = 0
        visitados  = set()
        wait_short = WebDriverWait(driver, 5)

        for i, card in enumerate(cards):
            if len(leads) >= max_leads:
                break
            try:
                href = card.get_attribute("href") or ""
                if href in visitados: continue
                visitados.add(href)

                log(f"🔍 [{i+1}/{len(cards)}] Abrindo empresa...")
                driver.execute_script("arguments[0].click();", card)
                time.sleep(3)

                lead = {
                    "nome": "", "nicho": nicho, "telefone": "", "email": "",
                    "instagram": "", "cidade": "", "ticket_medio": "",
                    "tempo_empresa": "", "porte": "", "cnpj": "", "info_empresa": "", "site": ""
                }

                # Nome
                for sel in ['h1[class*="fontHeadlineLarge"]', 'h1[class*="DUwDvf"]',
                            'h1.DUwDvf', 'div[role="main"] h1', 'h1']:
                    try:
                        el = wait_short.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
                        nome = el.text.strip()
                        if nome:
                            lead["nome"] = nome
                            break
                    except Exception:
                        pass

                if not lead["nome"]:
                    continue

                # Telefone
                try:
                    spans = driver.find_elements(By.XPATH,
                        '//button[@data-tooltip="Copiar número de telefone"]'
                        '//div[contains(@class,"fontBodyMedium")]')
                    if not spans:
                        spans = driver.find_elements(By.XPATH,
                            '//*[contains(@aria-label,"Ligue") or contains(@aria-label,"Phone")]')
                    for el in spans:
                        t = limpar_tel(el.get_attribute("aria-label") or el.text)
                        if re.search(r"\d{4,}", t):
                            lead["telefone"] = t; break
                except Exception:
                    pass

                if not lead["telefone"]:
                    try:
                        for btn in driver.find_elements(By.XPATH, '//button[@aria-label]'):
                            lbl = btn.get_attribute("aria-label") or ""
                            if re.search(r"\(\d{2}\)|\+\d{2}|\d{8,}", lbl):
                                lead["telefone"] = limpar_tel(lbl); break
                    except Exception:
                        pass

                # Site
                try:
                    sb = driver.find_element(By.XPATH,
                        '//a[@data-tooltip="Abrir site" or @data-item-id="authority"]')
                    lead["site"] = sb.get_attribute("href") or ""
                except Exception:
                    pass

                # Endereço / Cidade
                try:
                    for xpath in [
                        '//button[@data-item-id="address"]',
                        '//button[contains(@aria-label,"Endereço") or contains(@aria-label,"Address")]',
                    ]:
                        els = driver.find_elements(By.XPATH, xpath)
                        for el in els:
                            addr = (el.get_attribute("aria-label") or el.text or "").strip()
                            if addr:
                                lead["cidade"] = extrair_cidade(addr)
                                break
                        if lead["cidade"]: break
                except Exception:
                    pass

                # Ticket médio (faixa de preço do Google Maps)
                try:
                    for xpath in [
                        '//*[contains(@aria-label,"Nível de preço") or contains(@aria-label,"Price range")]',
                        '//span[contains(@aria-label,"€") or contains(@aria-label,"$") or contains(@aria-label,"R$")]',
                        '//button[contains(@jsaction,"pane.rating.moreReviews")]/..//span[contains(text(),"€") or contains(text(),"$$")]',
                    ]:
                        els = driver.find_elements(By.XPATH, xpath)
                        for el in els:
                            txt = (el.get_attribute("aria-label") or el.text or "").strip()
                            if txt:
                                lead["ticket_medio"] = txt; break
                        if lead["ticket_medio"]: break
                    # Fallback: texto com símbolos de preço visíveis no painel
                    if not lead["ticket_medio"]:
                        spans = driver.find_elements(By.XPATH,
                            '//span[contains(@class,"mgr77e") or contains(@class,"ZDu9vd")]')
                        for s in spans:
                            txt = s.text.strip()
                            if re.match(r'^[€$£R$·\s]+$', txt) and len(txt.replace('·','').strip()) > 0:
                                lead["ticket_medio"] = txt.replace('·', '').strip(); break
                except Exception:
                    pass

                # Info da empresa (descrição no Maps)
                try:
                    for xpath in [
                        '//div[contains(@class,"PYvSYb")]',
                        '//div[@aria-label="Sobre este lugar"]//span[not(contains(@class,"icon"))]',
                        '//h2[contains(text(),"Sobre")]/following-sibling::div//span',
                        '//div[contains(@class,"iP2t7d")]',
                    ]:
                        els = driver.find_elements(By.XPATH, xpath)
                        for el in els:
                            txt = el.text.strip()
                            if txt and len(txt) > 20:
                                lead["info_empresa"] = txt[:280]; break
                        if lead["info_empresa"]: break
                except Exception:
                    pass

                # Tempo de empresa (Maps: "X anos no Google" / "Aberto desde XXXX")
                try:
                    for xpath in [
                        '//*[contains(text(),"anos no Google") or contains(text(),"anos neste negócio")]',
                        '//*[contains(text(),"Aberto há") or contains(text(),"Estabelecido em")]',
                        '//*[contains(@aria-label,"anos")]',
                    ]:
                        els = driver.find_elements(By.XPATH, xpath)
                        for el in els:
                            txt = el.text.strip()
                            if txt and re.search(r'\d', txt):
                                lead["tempo_empresa"] = txt; break
                        if lead["tempo_empresa"]: break
                except Exception:
                    pass

                # Site + BrasilAPI CNPJ
                if lead["site"]:
                    log(f"   [{len(leads)+1}] {lead['nome'][:38]} — buscando site...")
                    sd = buscar_site(lead["site"])
                    lead["email"]     = sd["email"]
                    lead["instagram"] = sd["instagram"]
                    if not lead["tempo_empresa"] and sd["tempo"]:
                        lead["tempo_empresa"] = sd["tempo"]
                    if not lead["info_empresa"] and sd["info"]:
                        lead["info_empresa"] = sd["info"]

                    cnpj_data = sd.get("cnpj_data", {})
                    if cnpj_data:
                        lead["cnpj"] = cnpj_data.get("cnpj_fmt", "")
                        if cnpj_data.get("porte"):
                            lead["porte"] = cnpj_data["porte"]
                        # CNPJ tem precedência para tempo (data oficial da Receita)
                        if cnpj_data.get("tempo_empresa"):
                            lead["tempo_empresa"] = cnpj_data["tempo_empresa"]
                        # Cidade do CNPJ como fallback
                        if not lead["cidade"] and cnpj_data.get("cidade_cnpj"):
                            lead["cidade"] = cnpj_data["cidade_cnpj"]
                        # Atividade principal enriquece info
                        if cnpj_data.get("atividade"):
                            ativ = cnpj_data["atividade"]
                            if lead["info_empresa"]:
                                lead["info_empresa"] = f"[{ativ}] {lead['info_empresa']}"
                            else:
                                lead["info_empresa"] = ativ
                        if cnpj_data.get("cnpj_fmt"):
                            log(f"      🏢 CNPJ {cnpj_data['cnpj_fmt']} | {cnpj_data.get('porte','')} | {cnpj_data.get('tempo_empresa','')}")
                else:
                    log(f"   [{len(leads)+1}] {lead['nome'][:38]} — sem site")

                # DuckDuckGo: busca ativa quando ainda faltam dados
                falta_dados = not lead["cnpj"] or not lead["email"] or not lead["info_empresa"]
                if falta_dados:
                    cidade_busca = lead["cidade"] or local
                    log(f"      Buscando '{lead['nome'][:30]}' no DuckDuckGo...")
                    ddg = pesquisar_ddg(lead["nome"], cidade_busca)

                    if not lead["email"] and ddg["email"]:
                        lead["email"] = ddg["email"]
                    if not lead["instagram"] and ddg["instagram"]:
                        lead["instagram"] = ddg["instagram"]
                    if not lead["info_empresa"] and ddg["info"]:
                        lead["info_empresa"] = ddg["info"]

                    # CNPJ encontrado no DuckDuckGo → consulta BrasilAPI
                    if not lead["cnpj"] and ddg["cnpj"]:
                        log(f"      CNPJ {ddg['cnpj'][:14]}... encontrado via DDG")
                        cnpj_data = buscar_cnpj(ddg["cnpj"])
                        if cnpj_data:
                            lead["cnpj"] = cnpj_data.get("cnpj_fmt", "")
                            if cnpj_data.get("porte"):
                                lead["porte"] = cnpj_data["porte"]
                            if cnpj_data.get("tempo_empresa"):
                                lead["tempo_empresa"] = cnpj_data["tempo_empresa"]
                            if not lead["cidade"] and cnpj_data.get("cidade_cnpj"):
                                lead["cidade"] = cnpj_data["cidade_cnpj"]
                            if cnpj_data.get("atividade"):
                                ativ = cnpj_data["atividade"]
                                if lead["info_empresa"] and ativ not in lead["info_empresa"]:
                                    lead["info_empresa"] = f"[{ativ}] {lead['info_empresa']}"
                                elif not lead["info_empresa"]:
                                    lead["info_empresa"] = ativ

                chave = _chave_lead(lead)
                if chave in chaves_set:
                    pulados += 1
                    log(f"   ⏭️  {lead['nome'][:38]} — já coletado, pulando")
                    continue

                chaves_set.add(chave)
                leads.append(lead)
                log(f"   ✓ Tel:{lead['telefone'] or '—'} | Email:{lead['email'] or '—'} | "
                    f"Cidade:{lead['cidade'] or '—'} | Porte:{lead['porte'] or '—'} | "
                    f"Tempo:{lead['tempo_empresa'] or '—'}")
                result_q.put(("lead", lead))

            except Exception as e:
                log(f"   [!] Erro no card {i+1}: {e}")

        try:
            driver.quit()
        except Exception:
            pass

        hist["chaves"] = list(chaves_set)
        hist["total"]  = len(chaves_set)
        salvar_historico(hist)
        if pulados:
            log(f"💾 {pulados} empresa(s) ignoradas (já coletadas antes).")
        log(f"💾 Histórico: {len(chaves_set)} empresas únicas no total.")
        result_q.put(("done", leads))

    except Exception as e:
        log_q.put(f"❌ Erro fatal: {e}")
        result_q.put(("done", []))


# ── Form ───────────────────────────────────────────────────────────────────────
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("### Configurar busca")

col1, col2 = st.columns(2)
with col1:
    nicho = st.text_input("Nicho / Segmento", placeholder="Ex: clínica odontológica")
with col2:
    local = st.text_input("Cidade / Região", value="São Paulo", placeholder="Ex: Porto, Portugal")

num_leads = st.slider("Quantidade máxima de leads", min_value=5, max_value=200, value=30, step=5)
iniciar   = st.button("🚀 Iniciar coleta", disabled=st.session_state.running)
st.markdown('</div>', unsafe_allow_html=True)

# ── Execução ───────────────────────────────────────────────────────────────────
if iniciar:
    if not nicho.strip():
        st.error("Informe o nicho antes de iniciar.")
    else:
        st.session_state.update(running=True, leads=[], logs=[], done=False, excel_bytes=None)

        log_q    = queue.Queue()
        result_q = queue.Queue()

        threading.Thread(
            target=worker,
            args=(nicho.strip(), local.strip(), num_leads, log_q, result_q),
            daemon=True,
        ).start()

        prog_bar = st.progress(0, text="Iniciando…")
        log_ph   = st.empty()
        leads_ph = st.empty()

        while True:
            while not log_q.empty():
                st.session_state.logs.append(log_q.get())
            while not result_q.empty():
                kind, data = result_q.get()
                if kind == "lead":
                    st.session_state.leads.append(data)
                elif kind == "done":
                    st.session_state.done = True

            n   = len(st.session_state.leads)
            pct = min(int(n / max(num_leads, 1) * 100), 99)
            prog_bar.progress(pct, text=f"{n} lead(s) coletado(s)…")
            log_ph.markdown(
                f'<div class="log-box">{"<br>".join(st.session_state.logs[-60:])}</div>',
                unsafe_allow_html=True
            )
            if st.session_state.leads:
                leads_ph.markdown(_leads_html_table(st.session_state.leads), unsafe_allow_html=True)
            if st.session_state.done:
                break
            time.sleep(1)

        prog_bar.progress(100, text="Concluído!")
        st.session_state.running = False
        if st.session_state.leads:
            st.session_state.excel_bytes = gerar_excel(
                st.session_state.leads, nicho.strip(), local.strip()
            )

# ── Resultados ─────────────────────────────────────────────────────────────────
if st.session_state.leads:
    leads   = st.session_state.leads
    n_tel   = sum(1 for l in leads if l.get("telefone"))
    n_email = sum(1 for l in leads if l.get("email"))
    n_insta = sum(1 for l in leads if l.get("instagram"))
    n_cid   = sum(1 for l in leads if l.get("cidade"))
    n_info  = sum(1 for l in leads if l.get("info_empresa"))

    st.markdown("---")
    st.markdown("### Resultado")

    cols_stat = st.columns(5)
    for col, num, lbl in zip(
        cols_stat,
        [len(leads), n_tel, n_email, n_insta, n_cid],
        ["Empresas", "Com telefone", "Com e-mail", "Com Instagram", "Com cidade"],
    ):
        col.markdown(
            f'<div class="stat-box"><div class="num">{num}</div>'
            f'<div class="lbl">{lbl}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(_leads_html_table(leads), unsafe_allow_html=True)

    if not st.session_state.excel_bytes:
        st.session_state.excel_bytes = gerar_excel(leads, nicho.strip(), local.strip())

    if st.session_state.excel_bytes:
        st.download_button(
            label="⬇️  Baixar Excel",
            data=st.session_state.excel_bytes,
            file_name=f"leads_{nicho.replace(' ','_')}_{local.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

elif st.session_state.done and not st.session_state.leads:
    st.warning("Nenhum lead encontrado. Tente outro nicho ou localidade.")
