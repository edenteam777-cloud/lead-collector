#!/usr/bin/env python3
"""
Coletor de Leads - Geração automática de leads por nicho
Fonte: Google Maps
Saída: Nome da empresa, Email, Telefone → Excel
"""

import time
import re
import sys
import argparse
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, urlparse
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─── Configurações ────────────────────────────────────────────────────────────

SCROLL_PAUSE = 2.0          # segundos entre scrolls no painel de resultados
PAGE_LOAD_WAIT = 4          # segundos após clicar em um resultado
EMAIL_REQUEST_TIMEOUT = 8   # segundos para buscar email no site
MAX_EMAILS_PER_SITE = 1     # quantos emails capturar por empresa

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

# Domínios de email genéricos que devem ser ignorados
IGNORE_EMAIL_DOMAINS = {
    "exemplo.com", "example.com", "seudominio.com", "email.com",
    "seunome.com", "wixpress.com", "sentry.io", "suporte.com",
}


# ─── Utilitários ──────────────────────────────────────────────────────────────

def limpar_telefone(texto: str) -> str:
    return re.sub(r"[^\d+()\s\-]", "", texto).strip()


def formatar_url(url: str) -> str:
    if not url:
        return ""
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def buscar_emails_no_site(url: str) -> list[str]:
    """Visita o site da empresa e tenta extrair emails."""
    emails = set()
    url = formatar_url(url)
    if not url:
        return []

    try:
        resp = requests.get(
            url, headers=HEADERS, timeout=EMAIL_REQUEST_TIMEOUT,
            verify=False, allow_redirects=True
        )
        soup = BeautifulSoup(resp.text, "html.parser")

        # 1. Busca no texto da página
        for match in EMAIL_RE.findall(resp.text):
            dominio = match.split("@")[-1].lower()
            if dominio not in IGNORE_EMAIL_DOMAINS:
                emails.add(match.lower())
                if len(emails) >= MAX_EMAILS_PER_SITE:
                    return list(emails)

        # 2. Busca em links mailto:
        for tag in soup.find_all("a", href=True):
            href = tag["href"]
            if href.startswith("mailto:"):
                email = href.replace("mailto:", "").split("?")[0].strip().lower()
                dominio = email.split("@")[-1]
                if "@" in email and dominio not in IGNORE_EMAIL_DOMAINS:
                    emails.add(email)

        # 3. Tenta página /contato ou /contact
        if not emails:
            base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
            for slug in ["/contato", "/contact", "/fale-conosco", "/sobre"]:
                try:
                    r2 = requests.get(
                        base + slug, headers=HEADERS,
                        timeout=EMAIL_REQUEST_TIMEOUT, verify=False
                    )
                    for match in EMAIL_RE.findall(r2.text):
                        dominio = match.split("@")[-1].lower()
                        if dominio not in IGNORE_EMAIL_DOMAINS:
                            emails.add(match.lower())
                    if emails:
                        break
                except Exception:
                    pass

    except Exception:
        pass

    return list(emails)[:MAX_EMAILS_PER_SITE]


# ─── Scraping Google Maps ─────────────────────────────────────────────────────

def iniciar_driver() -> uc.Chrome:
    opts = uc.ChromeOptions()
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--lang=pt-BR")
    # Comente a linha abaixo se quiser ver o navegador
    opts.add_argument("--headless=new")
    driver = uc.Chrome(options=opts, version_main=146)
    driver.set_window_size(1400, 900)
    return driver


def scrollar_painel(driver, n_scrolls: int = 15):
    """Rola o painel lateral do Google Maps para carregar mais resultados."""
    try:
        painel = driver.find_element(
            By.XPATH,
            '//div[contains(@aria-label,"Resultados") or '
            'contains(@aria-label,"Results")][@role="feed"]'
        )
    except Exception:
        painel = None

    for _ in range(n_scrolls):
        try:
            if painel:
                driver.execute_script(
                    "arguments[0].scrollBy(0, 800);", painel
                )
            else:
                driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)
            time.sleep(SCROLL_PAUSE)
        except Exception:
            break


def coletar_cards(driver) -> list:
    """Retorna elementos de card de empresa no painel lateral."""
    seletores = [
        'a[href*="/maps/place/"]',
        'div[role="article"] a[href*="maps"]',
    ]
    for sel in seletores:
        cards = driver.find_elements(By.CSS_SELECTOR, sel)
        if cards:
            return cards
    return []


def extrair_detalhe(driver, wait) -> dict:
    """Extrai nome, telefone e site do painel de detalhe aberto."""
    lead = {"nome": "", "telefone": "", "site": "", "email": ""}

    try:
        # Nome
        nome_el = wait.until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'h1[class*="fontHeadlineLarge"]')
            )
        )
        lead["nome"] = nome_el.text.strip()
    except Exception:
        pass

    # Telefone — procura spans que pareçam número
    try:
        spans = driver.find_elements(
            By.XPATH,
            '//button[@data-tooltip="Copiar número de telefone"]'
            '//div[contains(@class,"fontBodyMedium")]'
        )
        if not spans:
            # fallback: qualquer span que contenha padrão telefônico
            spans = driver.find_elements(
                By.XPATH,
                '//*[contains(@aria-label,"Ligue") or '
                'contains(@aria-label,"Phone")]'
            )
        for el in spans:
            t = limpar_telefone(el.get_attribute("aria-label") or el.text)
            if re.search(r"\d{4,}", t):
                lead["telefone"] = t
                break
    except Exception:
        pass

    # Telefone — abordagem alternativa por aria-label
    if not lead["telefone"]:
        try:
            for btn in driver.find_elements(By.XPATH, '//button[@aria-label]'):
                lbl = btn.get_attribute("aria-label") or ""
                if re.search(r"\(\d{2}\)|\+\d{2}|\d{8,}", lbl):
                    lead["telefone"] = limpar_telefone(lbl)
                    break
        except Exception:
            pass

    # Site
    try:
        site_btn = driver.find_element(
            By.XPATH,
            '//a[@data-tooltip="Abrir site" or @data-item-id="authority"]'
        )
        lead["site"] = site_btn.get_attribute("href") or ""
    except Exception:
        pass

    return lead


def pesquisar_google_maps(nicho: str, local: str, max_leads: int, driver) -> list[dict]:
    query = f"{nicho} em {local}"
    url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
    print(f"\n[→] Abrindo Google Maps: {query}")
    driver.get(url)
    time.sleep(3)

    wait = WebDriverWait(driver, 12)

    # Quantos scrolls fazer (aprox. 5 resultados por scroll)
    n_scrolls = max(10, (max_leads // 5) + 5)
    print(f"[→] Carregando resultados (scrolls: {n_scrolls})...")
    scrollar_painel(driver, n_scrolls)

    cards = coletar_cards(driver)
    print(f"[→] {len(cards)} cards encontrados. Coletando detalhes...\n")

    leads = []
    visitados = set()

    for i, card in enumerate(cards):
        if len(leads) >= max_leads:
            break
        try:
            href = card.get_attribute("href") or ""
            if href in visitados:
                continue
            visitados.add(href)

            driver.execute_script("arguments[0].click();", card)
            time.sleep(PAGE_LOAD_WAIT)

            lead = extrair_detalhe(driver, wait)
            if not lead["nome"]:
                continue

            # Busca email no site
            if lead["site"]:
                print(f"   [{i+1}] {lead['nome']} — buscando email em {lead['site'][:50]}...")
                emails = buscar_emails_no_site(lead["site"])
                lead["email"] = emails[0] if emails else ""
            else:
                print(f"   [{i+1}] {lead['nome']} — sem site")

            leads.append(lead)
            print(
                f"       ✓ Tel: {lead['telefone'] or '—'} | "
                f"Email: {lead['email'] or '—'}"
            )

        except Exception as e:
            print(f"   [!] Erro no card {i+1}: {e}")
            continue

    return leads


# ─── Exportar Excel ───────────────────────────────────────────────────────────

COR_CABECALHO = "1F4E79"   # Azul escuro
COR_LINHA_PAR = "D9E1F2"   # Azul claro


def estilo_borda():
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_excel(leads: list[dict], nicho: str, local: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    colunas = ["#", "Nome da Empresa", "Telefone", "Email", "Site"]
    larguras = [5, 40, 20, 38, 45]

    # Cabeçalho
    for col, (titulo, larg) in enumerate(zip(colunas, larguras), start=1):
        cel = ws.cell(row=1, column=col, value=titulo)
        cel.font = Font(bold=True, color="FFFFFF", size=11)
        cel.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = estilo_borda()
        ws.column_dimensions[get_column_letter(col)].width = larg

    ws.row_dimensions[1].height = 22

    # Dados
    for idx, lead in enumerate(leads, start=1):
        row = idx + 1
        valores = [
            idx,
            lead.get("nome", ""),
            lead.get("telefone", ""),
            lead.get("email", ""),
            lead.get("site", ""),
        ]
        fill = PatternFill("solid", fgColor=COR_LINHA_PAR) if idx % 2 == 0 else None

        for col, valor in enumerate(valores, start=1):
            cel = ws.cell(row=row, column=col, value=valor)
            cel.alignment = Alignment(vertical="center", wrap_text=False)
            cel.border = estilo_borda()
            if fill:
                cel.fill = fill

    ws.freeze_panes = "A2"

    # Metadados na linha de título (acima do cabeçalho)
    ws.insert_rows(1)
    ws.merge_cells("A1:E1")
    meta = ws["A1"]
    meta.value = f"Leads — {nicho.title()} em {local.title()} | {len(leads)} empresas"
    meta.font = Font(bold=True, size=13, color="1F4E79")
    meta.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    nome_arquivo = f"leads_{nicho.replace(' ', '_')}_{local.replace(' ', '_')}.xlsx"
    wb.save(nome_arquivo)
    return nome_arquivo


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Gera leads (empresa, email, telefone) por nicho via Google Maps"
    )
    parser.add_argument("nicho", help='Nicho/segmento. Ex: "clínica odontológica"')
    parser.add_argument(
        "local",
        nargs="?",
        default="São Paulo",
        help='Cidade/região. Ex: "Curitiba, PR" (padrão: São Paulo)',
    )
    parser.add_argument(
        "-n", "--numero",
        type=int, default=30,
        help="Quantidade máxima de leads (padrão: 30)",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  COLETOR DE LEADS")
    print(f"  Nicho : {args.nicho}")
    print(f"  Local : {args.local}")
    print(f"  Meta  : {args.numero} leads")
    print("=" * 60)

    driver = iniciar_driver()
    try:
        leads = pesquisar_google_maps(args.nicho, args.local, args.numero, driver)
    finally:
        driver.quit()

    if not leads:
        print("\n[!] Nenhum lead encontrado. Tente outro nicho ou local.")
        sys.exit(1)

    arquivo = exportar_excel(leads, args.nicho, args.local)

    print(f"\n{'='*60}")
    print(f"  {len(leads)} leads salvos em: {arquivo}")
    print(f"{'='*60}\n")

    # Preview no terminal
    print(f"{'#':<4} {'Empresa':<35} {'Telefone':<18} {'Email'}")
    print("-" * 90)
    for i, l in enumerate(leads, 1):
        print(
            f"{i:<4} {l['nome'][:34]:<35} "
            f"{l['telefone'][:17]:<18} {l['email']}"
        )


if __name__ == "__main__":
    main()
